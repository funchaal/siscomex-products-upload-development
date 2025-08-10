import os

import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Side, Border, Alignment

import pandas as pd

from modules.utils import normalize_column_names

def makeFillSheet(products, relation_json):

    data = []

    possible_names = {
        'codigo': {
            'names': ['codigo', 'cod'],
            'obrigatorio': False
        },
        'versao': {
            'names': ['versao', 'versao (sistema)', 'versao produto', 'versao prod'],
            'obrigatorio': False
        },
        'ncm': {
            'names': ['ncm'],
            'obrigatorio': True
        },
        'descricao': {
            'names': ['descricao'],
            'obrigatorio': True
        },
        'denominacao': {
            'names': ['denominacao'],
            'obrigatorio': True
        },
        'cpfCnpjRaiz': {
            'names': ['raiz', 'cnpj', 'cpf/cnpj raiz', 'cpf/cnpj', 'cpf', 'cpfcnpjraiz'],
            'obrigatorio': True
        },
        'situacao': {
            'names': ['situacao'],
            'obrigatorio': False
        },
        'modalidade': {
            'names': ['modalidade'],
            'obrigatorio': False
        },
        'codigoInterno': {
            'names': ['codigo interno', 'cod interno', 'codigo produto', 'cod produto'],
            'obrigatorio': False
        }
    }

    products = normalize_column_names(products, possible_names)

    for _, row in products.iterrows():
        ncm = row.get('ncm')
        
        if not ncm: continue

        ncm = str(ncm).replace('.', '')
        ncm = ncm.zfill(8)

        raiz = str(row.get('cpfCnpjRaiz')).replace('.', '')[:8]
        raiz = raiz.zfill(8)

        codigo = row.get('codigo')
        descricao = row.get('descricao')
        denominacao = row.get('denominacao')
        codigoInterno = row.get('codigoInterno')
        modalidade = row.get('modalidade', 'IMPORTACAO')
        situacao = row.get('situacao', 'ATIVADO')

        base_info = {
            'Código': codigo,
            'NCM': ncm,
            'Raíz (CNPJ)': raiz,
            'Código Interno': codigoInterno,
            'Descrição': descricao,
            'Denominação': denominacao,
            'Modalidade': modalidade,
            'Situação': situacao, 
            'Atributos': []
        }

        def proccess_row(atribute_brute, type, multivalue_attr=False):

            atribute = None

            if type == 2:
                atribute = atribute_brute['atributo']
            else:
                atribute = atribute_brute

            attr_code = atribute['codigo']
            attr_name = atribute['nomeApresentacao']

            column_name = f'{attr_code} - {attr_name}'

            orientacaoPreenchimento = atribute.get('orientacaoPreenchimento', '')
            formaPreenchimento = atribute.get('formaPreenchimento', '')

            if orientacaoPreenchimento:
                column_name = column_name + '\n\n' + orientacaoPreenchimento

            if formaPreenchimento:
                column_name = column_name + '\n\nTipo preenchimento: ' + formaPreenchimento

            if type == 2:
                column_name = column_name + '\n\n(atributo condicional, preencha apenas se ' + atribute_brute['descricaoCondicao'] + ')'
            
            if multivalue_attr:
                column_name = column_name + '\n\nAtributo multivalorado, preencha os valores separados por vírgula'

            if formaPreenchimento == 'LISTA_ESTATICA':
                options = '\n\nPreencha um dos códigos abaixo:\n\n'
                for option in atribute['dominio']:
                    options = options + option['codigo'] + ' - ' + option['descricao'] + '\n'
                
                column_name = column_name + options

            column_name = str(column_name).replace('BOOLEANO', 'SIM ou NÃO').replace('true', 'SIM').replace('false', 'NÃO')

            value = None

            name = None

            for col in products.columns:
                parts = col.split('-')
                if not (len(parts) > 0 and 'ATT' in parts[0]):
                    parts = col.split('\n')
                    if len(parts) > 0:
                        name = parts[0].strip().lower()
                    else:
                        name = col

                if attr_code.lower() in parts[0].lower() or (name and attr_name.lower() == name):
                    value = row[col]

            # verificar se tem subatributos
            if len(atribute['listaSubatributos']) > 0:
                for sub_attr in atribute['listaSubatributos']:

                    field_type = None

                    if sub_attr['obrigatorio']:
                        field_type = 1
                    else:
                        field_type = 3

                    proccess_row(sub_attr, field_type)
                
                return
            
            base_info['Atributos'].append({
                'column name': column_name, 
                'field type': type, 
                'value': value
            })

            if atribute['atributoCondicionante']:
                for atribute_cond in atribute['condicionados']:
                    proccess_row(atribute_cond, 2, multivalue_attr=atribute_cond['atributo']['multivalorado'])

        attributes = next((a['listaAtributos'] for a in relation_json['listaNcm'] if str(a['codigoNcm']).replace('.', '') == str(ncm)), None)

        for attr in attributes:
            attr_dtls = next((d for d in relation_json['detalhesAtributos'] if d['codigo'] == attr['codigo']), None)

            field_type = None

            if attr['obrigatorio']:
                field_type = 1
            else:
                field_type = 3

            proccess_row(attr_dtls, field_type, multivalue_attr=attr['multivalorado'])

        data.append(base_info)

    wb = convert_to_excel(data)

    return wb

def convert_to_excel(data):
    # Identificar todas as colunas fixas e dinâmicas
    fixed_columns = ["Código", "NCM", "Raíz (CNPJ)", "Descrição", "Denominação", "Código Interno", "Modalidade", "Situação"]
    attribute_columns = []
    
    for row in data:
        for attr in row.get("Atributos", []):
            if attr["column name"] not in attribute_columns:
                attribute_columns.append(attr["column name"])
    
    all_columns = fixed_columns + attribute_columns
    
    # Criar DataFrame vazio
    df = pd.DataFrame(columns=all_columns)
    
    # Popular DataFrame
    records = []
    attribute_styles = {}
    
    for row in data:
        record = {col: row.get(col, "") for col in fixed_columns}
        row_attributes = {attr["column name"]: attr for attr in row.get("Atributos", [])}
        
        for col_name in attribute_columns:
            if col_name in row_attributes:
                attr = row_attributes[col_name]
                record[col_name] = attr["value"]
                attribute_styles[(len(records) + 1, all_columns.index(col_name) + 1)] = attr["field type"]
            else:
                record[col_name] = ""
                attribute_styles[(len(records) + 1, all_columns.index(col_name) + 1)] = 0  # Vermelho para ausência do atributo
        
        records.append(record)
    
    # Criar um novo Workbook e preencher com os dados do DataFrame
    wb = Workbook()
    ws = wb.active
    ws.title = "Preenchimento"
    ws.append(all_columns)
    
    for row in records:
        ws.append([row.get(col, "") for col in all_columns])
    
    color_mapping = {
        1: "FFFFFF",  # Branco
        2: "FFFF00",  # Amarelo
        3: "D3D3D3",  # Cinza Claro
        0: "FF0000"   # Vermelho para ausência do atributo
    }

    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for (row, col), field_type in attribute_styles.items():
        cell = ws.cell(row=row+1, column=col)
        cell.fill = PatternFill(start_color=color_mapping.get(field_type, "FFFFFF"), fill_type="solid")
        cell.border = border_style

    def formatar_cabecalho(ws):
        header_row = ws[1]
        for cell in header_row:
            # Centraliza o texto e permite quebra de linha
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Ajusta a largura da coluna baseado no número da coluna da célula
            col_letter = get_column_letter(cell.column)  # Obter a letra da coluna
            ws.column_dimensions[col_letter].width = 40  # Ajuste conforme necessário

        # Ajuste de altura da linha 1
        ws.row_dimensions[1].height = 400  # Ajuste conforme o texto ou o conteúdo da célula

    formatar_cabecalho(ws)

    adicionar_aba_legenda(wb, border_style)
    
    return wb

def adicionar_aba_legenda(wb, border_style):
    ws = wb.create_sheet("Legenda")

    # Cabeçalho com apenas "Cor" e "Legenda"
    ws.append(["Cor", "Legenda"])

    # Mapeamento de cor para descrição
    legenda_cores = {
        "FFFFFF": "Branco",
        "FFFF00": "Amarelo",
        "D3D3D3": "Cinza",
        "FF0000": "Vermelho"
    }

    significado = {
        "FFFFFF": "Obrigatório",
        "FFFF00": "Condicional",
        "D3D3D3": "Não obrigatório",
        "FF0000": "Não deve ser preenchido"
    }

    # Lista apenas com as cores desejadas
    cores = ["FFFFFF", "FFFF00", "D3D3D3", "FF0000"]

    for cor_hex in cores:
        ws.append([legenda_cores[cor_hex], significado[cor_hex]])
        row_idx = ws.max_row

        # Aplica cor de fundo na célula da coluna "Cor" (agora é a coluna 1)
        cor_cell = ws.cell(row=row_idx, column=1)
        cor_cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
        cor_cell.border = border_style

        # Aplica borda na coluna "Legenda" (agora é a coluna 2)
        ws.cell(row=row_idx, column=2).border = border_style

    # Borda no cabeçalho
    for col in range(1, 3):
        ws.cell(row=1, column=col).border = border_style

    # Ajuste de largura (opcional)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30